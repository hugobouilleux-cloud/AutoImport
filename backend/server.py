from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Form
from fastapi.responses import FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict
import uuid
from datetime import datetime, timezone
import httpx
from playwright.async_api import async_playwright
import asyncio
import json
from openpyxl import load_workbook


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")


# Define Models
class ConnectionConfig(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    site_url: str
    login: str
    password: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    last_connected: Optional[datetime] = None
    status: str = "pending"  # pending, success, failed

class ConnectionConfigCreate(BaseModel):
    site_url: str
    login: str
    password: str

class ConnectionTest(BaseModel):
    site_url: str
    login: str
    password: str

class ConnectionResult(BaseModel):
    success: bool
    message: str
    status_code: Optional[int] = None

class ImportFormat(BaseModel):
    name: str
    href: str

class ImportFormatsList(BaseModel):
    success: bool
    message: str
    formats: List[ImportFormat]
    total_count: int

class SelectFormatRequest(BaseModel):
    site_url: str
    login: str
    password: str
    selected_format: ImportFormat

class SelectFormatResult(BaseModel):
    success: bool
    message: str
    format_url: Optional[str] = None

class TableRow(BaseModel):
    cells: List[str]

class TableExtractionResult(BaseModel):
    success: bool
    message: str
    headers: List[str]
    rows: List[TableRow]
    total_rows: int

class FetchListsRequest(BaseModel):
    site_url: str
    login: str
    system_password: str
    table_config: TableExtractionResult

class ListFieldInfo(BaseModel):
    field_path: str
    list_type: str
    values: List[str]

class FetchListsResult(BaseModel):
    success: bool
    message: str
    list_fields: List[ListFieldInfo] = []

# Add your routes to the router instead of directly to app
@api_router.get("/")
async def root():
    return {"message": "Automation Import App API"}

@api_router.post("/connection/test", response_model=ConnectionResult)
async def test_connection(connection_data: ConnectionTest):
    """
    Test the connection to the external site
    """
    try:
        async with httpx.AsyncClient(follow_redirects=True, timeout=30.0, verify=False) as client:
            # Préparer les données de connexion
            login_data = {
                "username": connection_data.login,
                "password": connection_data.password
            }
            
            # Tenter la connexion
            response = await client.post(
                connection_data.site_url,
                data=login_data,
                headers={
                    "Content-Type": "application/x-www-form-urlencoded",
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
                }
            )
            
            # Vérifier la réponse
            if response.status_code == 200:
                # Vérifier si la connexion a réussi (peut nécessiter une adaptation selon le site)
                if "error" not in response.text.lower() and "invalid" not in response.text.lower():
                    return ConnectionResult(
                        success=True,
                        message="Connexion réussie au site cible",
                        status_code=response.status_code
                    )
                else:
                    return ConnectionResult(
                        success=False,
                        message="Identifiants invalides",
                        status_code=response.status_code
                    )
            else:
                return ConnectionResult(
                    success=False,
                    message=f"Échec de la connexion: HTTP {response.status_code}",
                    status_code=response.status_code
                )
                
    except httpx.TimeoutException:
        return ConnectionResult(
            success=False,
            message="Timeout: le site ne répond pas",
            status_code=None
        )
    except httpx.RequestError as e:
        return ConnectionResult(
            success=False,
            message=f"Erreur de connexion: {str(e)}",
            status_code=None
        )
    except Exception as e:
        return ConnectionResult(
            success=False,
            message=f"Erreur inattendue: {str(e)}",
            status_code=None
        )

@api_router.post("/connection/save", response_model=ConnectionConfig)
async def save_connection(connection_data: ConnectionConfigCreate):
    """
    Save connection configuration
    """
    config = ConnectionConfig(**connection_data.model_dump())
    
    # Convert to dict and serialize datetime to ISO string for MongoDB
    doc = config.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    if doc['last_connected']:
        doc['last_connected'] = doc['last_connected'].isoformat()
    
    await db.connections.insert_one(doc)
    return config

@api_router.get("/connection/list", response_model=List[ConnectionConfig])
async def list_connections():
    """
    Get all saved connections
    """
    connections = await db.connections.find({}, {"_id": 0}).to_list(1000)
    
    # Convert ISO string timestamps back to datetime objects
    for conn in connections:
        if isinstance(conn['created_at'], str):
            conn['created_at'] = datetime.fromisoformat(conn['created_at'])
        if conn.get('last_connected') and isinstance(conn['last_connected'], str):
            conn['last_connected'] = datetime.fromisoformat(conn['last_connected'])
    
    return connections

@api_router.delete("/connection/{connection_id}")
async def delete_connection(connection_id: str):
    """
    Delete a saved connection
    """
    result = await db.connections.delete_one({"id": connection_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Connection not found")
    return {"message": "Connection deleted successfully"}

@api_router.post("/connection/navigate-admin")
async def navigate_to_admin(connection_data: ConnectionTest):
    """
    Navigate to administration page using Playwright
    """
    try:
        async with async_playwright() as p:
            # Lancer le navigateur en mode headless
            browser = await p.chromium.launch(headless=True, args=['--no-sandbox', '--disable-setuid-sandbox'])
            context = await browser.new_context(
                ignore_https_errors=True,
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            )
            page = await context.new_page()
            
            try:
                # Étape 1: Aller sur la page de login
                await page.goto(connection_data.site_url, timeout=30000, wait_until="load")
                await asyncio.sleep(2)
                
                # Check if already logged in (look for user icon)
                already_logged_in = await page.query_selector('.icon-user')
                
                if not already_logged_in:
                    # Need to log in
                    logger.info("Connexion requise...")
                    
                    # Check if login form exists
                    login_form = await page.query_selector('input[name="j_username"]')
                    if not login_form:
                        await browser.close()
                        return {
                            "success": False,
                            "message": "Formulaire de connexion non trouvé"
                        }
                    
                    # Étape 2: Remplir le formulaire de connexion
                    await page.fill('input[name="j_username"]', connection_data.login, timeout=5000)
                    await asyncio.sleep(0.5)
                    await page.fill('input[name="j_password"]', connection_data.password, timeout=5000)
                    await asyncio.sleep(1)
                    
                    # Étape 3: Attendre que le bouton soit activé et cliquer
                    if not await click_login_button(page):
                        await browser.close()
                        return {
                            "success": False,
                            "message": "Bouton de connexion non trouvé"
                        }
                    
                    # Attendre la navigation
                    await page.wait_for_load_state("load", timeout=30000)
                    await asyncio.sleep(3)
                else:
                    logger.info("Déjà connecté, pas besoin de se reconnecter")
                
                # Étape 4: Cliquer sur l'icône utilisateur
                try:
                    # Chercher l'élément avec la classe icon-user
                    await page.click('.icon-user', timeout=5000)
                    await asyncio.sleep(1)
                except Exception as e:
                    # Essayer d'autres sélecteurs
                    user_selectors = [
                        'span.icon-user-without-picture',
                        'span.icon-user',
                        '[class*="icon-user"]',
                        'button:has(.icon-user)'
                    ]
                    clicked = False
                    for selector in user_selectors:
                        try:
                            await page.click(selector, timeout=2000)
                            clicked = True
                            await asyncio.sleep(1)
                            break
                        except:
                            continue
                    
                    if not clicked:
                        raise Exception(f"Impossible de cliquer sur l'icône utilisateur: {str(e)}")
                
                # Étape 5: Cliquer sur "Administration" dans le menu
                try:
                    # Attendre que le menu apparaisse
                    await page.wait_for_selector('button[mat-menu-item]', timeout=5000)
                    
                    # Cliquer sur Administration
                    admin_selectors = [
                        'button.user-menu-item:has-text("Administration")',
                        'button[mat-menu-item]:has-text("Administration")',
                        'span.user-menu-item-label:has-text("Administration")',
                        '[class*="menu-item"]:has-text("Administration")'
                    ]
                    
                    clicked_admin = False
                    for selector in admin_selectors:
                        try:
                            await page.click(selector, timeout=2000)
                            clicked_admin = True
                            break
                        except:
                            continue
                    
                    if not clicked_admin:
                        raise Exception("Impossible de cliquer sur Administration")
                    
                    # Attendre le chargement de la page d'administration
                    await page.wait_for_load_state("load", timeout=30000)
                    await asyncio.sleep(3)
                    
                    # Récupérer l'URL actuelle
                    current_url = page.url
                    
                    # Prendre un screenshot pour vérification
                    screenshot = await page.screenshot(full_page=False)
                    
                    await browser.close()
                    
                    return {
                        "success": True,
                        "message": "Navigation vers l'administration réussie",
                        "admin_url": current_url
                    }
                    
                except Exception as e:
                    await browser.close()
                    raise Exception(f"Erreur lors de la navigation vers l'administration: {str(e)}")
                
            except Exception as e:
                await browser.close()
                raise e
                
    except Exception as e:
        logger.error(f"Navigation error: {str(e)}")
        return {
            "success": False,
            "message": f"Erreur: {str(e)}",
            "admin_url": None
        }

@api_router.post("/connection/extract-formats", response_model=ImportFormatsList)
async def extract_import_formats(connection_data: ConnectionTest):
    """
    Extract all import formats from the admin page with pagination
    """
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True, args=['--no-sandbox', '--disable-setuid-sandbox'])
            context = await browser.new_context(
                ignore_https_errors=True,
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            )
            page = await context.new_page()
            
            try:
                # Étape 1: Connexion (même processus que navigate-admin)
                await page.goto(connection_data.site_url, timeout=30000, wait_until="load")
                await asyncio.sleep(1)
                
                await page.fill('input[name="j_username"]', connection_data.login, timeout=5000)
                await asyncio.sleep(0.5)
                await page.fill('input[name="j_password"]', connection_data.password, timeout=5000)
                await asyncio.sleep(1)
                
                if not await click_login_button(page):
                    await browser.close()
                    return {
                        "success": False,
                        "message": "Bouton de connexion non trouvé",
                        "formats": []
                    }
                
                # Wait for login to complete - use 'load' instead of 'networkidle'
                await page.wait_for_load_state("load", timeout=30000)
                await asyncio.sleep(2)
                
                # Wait for user icon to be available
                await page.wait_for_selector('.icon-user', timeout=10000)
                await page.click('.icon-user', timeout=5000)
                await asyncio.sleep(1)
                
                # Cliquer sur Administration
                await page.wait_for_selector('button[mat-menu-item]', timeout=10000)
                await page.click('button.user-menu-item:has-text("Administration")', timeout=5000)
                await page.wait_for_load_state("load", timeout=30000)
                await asyncio.sleep(3)
                
                # Étape 2: Cliquer sur "Import de données"
                await page.wait_for_selector('a:has-text("Import de données")', timeout=10000)
                await page.click('a:has-text("Import de données")', timeout=10000)
                await page.wait_for_load_state("load", timeout=30000)
                await asyncio.sleep(3)
                
                # Étape 3: Extraire tous les formats avec pagination
                all_formats = []
                page_number = 1
                
                while True:
                    # Attendre que le tableau soit chargé
                    await page.wait_for_selector('app-link a.a', timeout=10000)
                    await asyncio.sleep(1.5)
                    
                    # Extraire les éléments de la page actuelle
                    formats = await page.evaluate('''() => {
                        const links = document.querySelectorAll('app-link a.a');
                        return Array.from(links).map(link => ({
                            name: link.textContent.trim(),
                            href: link.getAttribute('href')
                        }));
                    }''')
                    
                    all_formats.extend(formats)
                    logger.info(f"Page {page_number}: {len(formats)} formats extraits, total: {len(all_formats)}")
                    
                    # Vérifier s'il y a une page suivante (bouton non désactivé)
                    try:
                        # Chercher le bouton "Page suivante" qui n'est pas disabled
                        next_button_exists = await page.evaluate('''() => {
                            const buttons = document.querySelectorAll('button.k-pager-nav');
                            for (let btn of buttons) {
                                const title = btn.getAttribute('title');
                                const ariaLabel = btn.getAttribute('aria-label');
                                if ((title === 'Page suivante' || ariaLabel === 'Page suivante') && 
                                    !btn.disabled && !btn.classList.contains('k-disabled')) {
                                    return true;
                                }
                            }
                            return false;
                        }''')
                        
                        if next_button_exists:
                            # Cliquer sur le bouton suivant
                            await page.evaluate('''() => {
                                const buttons = document.querySelectorAll('button.k-pager-nav');
                                for (let btn of buttons) {
                                    const title = btn.getAttribute('title');
                                    const ariaLabel = btn.getAttribute('aria-label');
                                    if ((title === 'Page suivante' || ariaLabel === 'Page suivante') && 
                                        !btn.disabled && !btn.classList.contains('k-disabled')) {
                                        btn.click();
                                        return;
                                    }
                                }
                            }''')
                            await asyncio.sleep(2)
                            await page.wait_for_load_state("load", timeout=20000)
                            await asyncio.sleep(2)
                            page_number += 1
                        else:
                            logger.info(f"Fin de la pagination - Total: {len(all_formats)} formats")
                            break
                    except Exception as e:
                        logger.info(f"Erreur pagination ou fin atteinte: {str(e)}")
                        break
                
                await browser.close()
                
                return ImportFormatsList(
                    success=True,
                    message=f"{len(all_formats)} formats d'import extraits avec succès",
                    formats=all_formats,
                    total_count=len(all_formats)
                )
                
            except Exception as e:
                await browser.close()
                raise e
                
    except Exception as e:
        logger.error(f"Extract formats error: {str(e)}")
        return ImportFormatsList(
            success=False,
            message=f"Erreur: {str(e)}",
            formats=[],
            total_count=0
        )

@api_router.post("/connection/select-format", response_model=SelectFormatResult)
async def select_format_in_table(request: SelectFormatRequest):
    """
    Navigate to the format page and click on the selected format in the table
    """
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True, args=['--no-sandbox', '--disable-setuid-sandbox'])
            context = await browser.new_context(
                ignore_https_errors=True,
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            )
            page = await context.new_page()
            
            try:
                # Étape 1: Connexion
                await page.goto(request.site_url, timeout=30000, wait_until="load")
                await asyncio.sleep(1)
                
                await page.fill('input[name="j_username"]', request.login, timeout=5000)
                await asyncio.sleep(0.5)
                await page.fill('input[name="j_password"]', request.password, timeout=5000)
                await asyncio.sleep(1)
                
                await page.wait_for_selector('button:has-text("Connexion")', timeout=10000)
                await asyncio.sleep(0.5)
                await page.click('button:has-text("Connexion")', timeout=5000)
                
                # Wait for login to complete - use 'load' instead of 'networkidle'
                await page.wait_for_load_state("load", timeout=30000)
                await asyncio.sleep(2)
                
                # Wait for user icon to be available
                await page.wait_for_selector('.icon-user', timeout=10000)
                await page.click('.icon-user', timeout=5000)
                await asyncio.sleep(1)
                
                # Cliquer sur Administration
                await page.wait_for_selector('button[mat-menu-item]', timeout=10000)
                await page.click('button.user-menu-item:has-text("Administration")', timeout=5000)
                await page.wait_for_load_state("load", timeout=30000)
                await asyncio.sleep(3)
                
                # Étape 2: Cliquer sur "Import de données"
                await page.wait_for_selector('a:has-text("Import de données")', timeout=10000)
                await page.click('a:has-text("Import de données")', timeout=10000)
                await page.wait_for_load_state("load", timeout=30000)
                await asyncio.sleep(3)
                
                # Étape 3: Chercher et cliquer sur le format sélectionné dans le tableau
                format_name = request.selected_format.name
                format_href = request.selected_format.href
                
                logger.info(f"Recherche du format: {format_name}")
                
                # Parcourir les pages pour trouver le format
                format_found = False
                page_number = 1
                
                while not format_found:
                    await page.wait_for_selector('app-link a.a', timeout=10000)
                    await asyncio.sleep(1.5)
                    
                    # Chercher le format sur la page actuelle
                    format_clicked = await page.evaluate('''(formatName) => {
                        const links = document.querySelectorAll('app-link a.a');
                        for (let link of links) {
                            if (link.textContent.trim() === formatName) {
                                link.click();
                                return true;
                            }
                        }
                        return false;
                    }''', format_name)
                    
                    if format_clicked:
                        format_found = True
                        logger.info(f"Format trouvé et cliqué sur la page {page_number}")
                        await asyncio.sleep(2)
                        await page.wait_for_load_state("load", timeout=30000)
                        
                        current_url = page.url
                        await browser.close()
                        
                        return SelectFormatResult(
                            success=True,
                            message=f"Format '{format_name}' sélectionné avec succès",
                            format_url=current_url
                        )
                    
                    # Si pas trouvé, aller à la page suivante
                    next_button_exists = await page.evaluate('''() => {
                        const buttons = document.querySelectorAll('button.k-pager-nav');
                        for (let btn of buttons) {
                            const title = btn.getAttribute('title');
                            const ariaLabel = btn.getAttribute('aria-label');
                            if ((title === 'Page suivante' || ariaLabel === 'Page suivante') && 
                                !btn.disabled && !btn.classList.contains('k-disabled')) {
                                return true;
                            }
                        }
                        return false;
                    }''')
                    
                    if next_button_exists:
                        await page.evaluate('''() => {
                            const buttons = document.querySelectorAll('button.k-pager-nav');
                            for (let btn of buttons) {
                                const title = btn.getAttribute('title');
                                const ariaLabel = btn.getAttribute('aria-label');
                                if ((title === 'Page suivante' || ariaLabel === 'Page suivante') && 
                                    !btn.disabled && !btn.classList.contains('k-disabled')) {
                                    btn.click();
                                    return;
                                }
                            }
                        }''')
                        await asyncio.sleep(2)
                        await page.wait_for_load_state("load", timeout=20000)
                        page_number += 1
                    else:
                        await browser.close()
                        return SelectFormatResult(
                            success=False,
                            message=f"Format '{format_name}' introuvable dans le tableau",
                            format_url=None
                        )
                
            except Exception as e:
                await browser.close()
                raise e
                
    except Exception as e:
        logger.error(f"Select format error: {str(e)}")
        return SelectFormatResult(
            success=False,
            message=f"Erreur: {str(e)}",
            format_url=None
        )

@api_router.post("/connection/extract-table", response_model=TableExtractionResult)
async def extract_format_table(request: SelectFormatRequest):
    """
    Extract the configuration table after selecting a format
    """
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True, args=['--no-sandbox', '--disable-setuid-sandbox'])
            context = await browser.new_context(
                ignore_https_errors=True,
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            )
            page = await context.new_page()
            
            try:
                # Étape 1: Connexion
                await page.goto(request.site_url, timeout=30000, wait_until="load")
                await asyncio.sleep(1)
                
                await page.fill('input[name="j_username"]', request.login, timeout=5000)
                await asyncio.sleep(0.5)
                await page.fill('input[name="j_password"]', request.password, timeout=5000)
                await asyncio.sleep(1)
                
                await page.wait_for_selector('button:has-text("Connexion")', timeout=10000)
                await asyncio.sleep(0.5)
                await page.click('button:has-text("Connexion")', timeout=5000)
                
                # Wait for login to complete - use 'load' instead of 'networkidle'
                await page.wait_for_load_state("load", timeout=30000)
                await asyncio.sleep(2)
                
                # Wait for user icon to be available
                await page.wait_for_selector('.icon-user', timeout=10000)
                await page.click('.icon-user', timeout=5000)
                await asyncio.sleep(1)
                
                # Cliquer sur Administration
                await page.wait_for_selector('button[mat-menu-item]', timeout=10000)
                await page.click('button.user-menu-item:has-text("Administration")', timeout=5000)
                await page.wait_for_load_state("load", timeout=30000)
                await asyncio.sleep(3)
                
                # Cliquer sur "Import de données"
                await page.wait_for_selector('a:has-text("Import de données")', timeout=10000)
                await page.click('a:has-text("Import de données")', timeout=10000)
                await page.wait_for_load_state("load", timeout=30000)
                await asyncio.sleep(3)
                
                # Chercher et cliquer sur le format
                format_name = request.selected_format.name
                page_number = 1
                format_found = False
                
                while not format_found:
                    await page.wait_for_selector('app-link a.a', timeout=10000)
                    await asyncio.sleep(1.5)
                    
                    logger.info(f"Recherche du format sur la page {page_number}...")
                    format_clicked = await page.evaluate('''(formatName) => {
                        const links = document.querySelectorAll('app-link a.a');
                        for (let link of links) {
                            if (link.textContent.trim() === formatName) {
                                console.log('Format trouvé et cliqué:', formatName);
                                link.click();
                                return true;
                            }
                        }
                        return false;
                    }''', format_name)
                    
                    if format_clicked:
                        format_found = True
                        logger.info(f"Format cliqué, attente du chargement...")
                        
                        # Attendre que l'URL change ou qu'un élément spécifique apparaisse
                        await asyncio.sleep(3)
                        
                        # Au lieu d'attendre networkidle, attendre un élément spécifique
                        try:
                            # Attendre qu'un élément de la nouvelle page soit présent
                            await page.wait_for_selector('kendo-grid, table.k-grid-table', timeout=30000)
                            logger.info("Nouvelle page détectée")
                        except Exception as e:
                            logger.warning(f"Timeout attente élément: {str(e)}")
                        
                        await asyncio.sleep(2)
                        break
                    
                    # Page suivante
                    next_button_exists = await page.evaluate('''() => {
                        const buttons = document.querySelectorAll('button.k-pager-nav');
                        for (let btn of buttons) {
                            const title = btn.getAttribute('title');
                            const ariaLabel = btn.getAttribute('aria-label');
                            if ((title === 'Page suivante' || ariaLabel === 'Page suivante') && 
                                !btn.disabled && !btn.classList.contains('k-disabled')) {
                                return true;
                            }
                        }
                        return false;
                    }''')
                    
                    if next_button_exists:
                        await page.evaluate('''() => {
                            const buttons = document.querySelectorAll('button.k-pager-nav');
                            for (let btn of buttons) {
                                const title = btn.getAttribute('title');
                                const ariaLabel = btn.getAttribute('aria-label');
                                if ((title === 'Page suivante' || ariaLabel === 'Page suivante') && 
                                    !btn.disabled && !btn.classList.contains('k-disabled')) {
                                    btn.click();
                                    return;
                                }
                            }
                        }''')
                        await asyncio.sleep(2)
                        await page.wait_for_load_state("load", timeout=20000)
                        page_number += 1
                    else:
                        await browser.close()
                        return TableExtractionResult(
                            success=False,
                            message=f"Format '{format_name}' introuvable",
                            headers=[],
                            rows=[],
                            total_rows=0
                        )
                
                # Extraire le tableau
                logger.info("Recherche du tableau de configuration...")
                
                # Attendre le tableau avec plusieurs tentatives
                table_found = False
                max_attempts = 3
                
                for attempt in range(max_attempts):
                    try:
                        logger.info(f"Tentative {attempt + 1}/{max_attempts}...")
                        await page.wait_for_selector('table.k-grid-table', timeout=15000)
                        table_found = True
                        logger.info("Tableau trouvé !")
                        break
                    except Exception as e:
                        logger.warning(f"Tentative {attempt + 1} échouée: {str(e)}")
                        if attempt < max_attempts - 1:
                            await asyncio.sleep(2)
                
                if not table_found:
                    # Dernier essai avec un sélecteur plus général
                    try:
                        await page.wait_for_selector('kendo-grid', timeout=10000)
                        logger.info("Kendo-grid trouvé")
                    except:
                        await browser.close()
                        return TableExtractionResult(
                            success=False,
                            message="Impossible de trouver le tableau de configuration",
                            headers=[],
                            rows=[],
                            total_rows=0
                        )
                
                await asyncio.sleep(2)
                
                table_data = await page.evaluate('''() => {
                    // Extraire les en-têtes
                    const headers = [];
                    const headerCells = document.querySelectorAll('thead.k-table-thead th.k-header span.k-column-title');
                    headerCells.forEach(cell => {
                        headers.push(cell.textContent.trim());
                    });
                    
                    // Extraire les lignes
                    const rows = [];
                    const bodyRows = document.querySelectorAll('tbody.k-table-tbody tr.k-table-row');
                    bodyRows.forEach(row => {
                        const cells = [];
                        const tds = row.querySelectorAll('td.k-table-td');
                        tds.forEach(td => {
                            // Extraire le texte ou vérifier les indicateurs bool
                            const span = td.querySelector('span:not(.bool)');
                            const boolIndicator = td.querySelector('span.bool');
                            
                            if (span && span.textContent.trim() !== '&nbsp;' && span.textContent.trim() !== '') {
                                cells.push(span.textContent.trim());
                            } else if (boolIndicator) {
                                // val-1 = true (checked), val-2 = false (unchecked)
                                const isChecked = boolIndicator.classList.contains('val-1');
                                cells.push(isChecked ? 'Oui' : 'Non');
                            } else {
                                cells.push('');
                            }
                        });
                        if (cells.length > 0) {
                            rows.push({ cells: cells });
                        }
                    });
                    
                    return { headers, rows };
                }''')
                
                await browser.close()
                
                logger.info(f"Table extracted: {len(table_data['headers'])} columns, {len(table_data['rows'])} rows")
                
                return TableExtractionResult(
                    success=True,
                    message=f"Tableau extrait avec succès: {len(table_data['rows'])} lignes",
                    headers=table_data['headers'],
                    rows=table_data['rows'],
                    total_rows=len(table_data['rows'])
                )
                
            except Exception as e:
                await browser.close()
                raise e
                
    except Exception as e:
        logger.error(f"Extract table error: {str(e)}")
        return TableExtractionResult(
            success=False,
            message=f"Erreur: {str(e)}",
            headers=[],
            rows=[],
            total_rows=0
        )

@api_router.post("/connection/fetch-lists", response_model=FetchListsResult)
async def fetch_reference_lists(request: FetchListsRequest):
    """
    Fetch reference list values from Legisway API after table extraction
    This allows users to see valid values before uploading their file
    """
    try:
        logger.info("Extraction des champs avec listes de référence...")
        
        # Extract fields with list filters from table_config
        list_fields_info = []
        table_config = request.table_config
        
        for row in table_config.rows:
            cells = row.cells if hasattr(row, 'cells') else row.get('cells', [])
            
            if len(cells) >= 3:
                field_path = cells[0]  # Chemin (path)
                filter_value = cells[2]  # Filtre
                
                # Check if filter contains type.name='...'
                if filter_value and "type.name=" in filter_value:
                    import re
                    match = re.search(r"type\.name\s*=\s*['\"]([^'\"]+)['\"]", filter_value)
                    if match:
                        list_type = match.group(1)
                        list_fields_info.append({
                            "field_path": field_path,
                            "list_type": list_type
                        })
        
        logger.info(f"Champs avec listes trouvés: {len(list_fields_info)}")
        
        if not list_fields_info:
            return FetchListsResult(
                success=True,
                message="Aucune liste de référence à récupérer",
                list_fields=[]
            )
        
        # Fetch list values from Legisway API
        list_types = [field['list_type'] for field in list_fields_info]
        logger.info(f"Récupération des listes: {list_types}")
        
        list_values_response = await fetch_list_values_from_legisway(
            site_url=request.site_url,
            login=request.login,
            system_password=request.system_password,
            list_types=list_types
        )
        
        if not list_values_response['success']:
            return FetchListsResult(
                success=False,
                message=f"Erreur récupération des listes: {list_values_response.get('message', 'Unknown error')}",
                list_fields=[]
            )
        
        # Build response with field info and values
        result_list_fields = []
        for field_info in list_fields_info:
            list_type = field_info['list_type']
            values = list_values_response['lists'].get(list_type, [])
            
            result_list_fields.append(ListFieldInfo(
                field_path=field_info['field_path'],
                list_type=list_type,
                values=values
            ))
        
        total_values = sum(len(field.values) for field in result_list_fields)
        
        return FetchListsResult(
            success=True,
            message=f"{len(result_list_fields)} listes récupérées ({total_values} valeurs au total)",
            list_fields=result_list_fields
        )
        
    except Exception as e:
        logger.error(f"Fetch lists error: {str(e)}")
        return FetchListsResult(
            success=False,
            message=f"Erreur: {str(e)}",
            list_fields=[]
        )

@api_router.post("/import/execute")
async def execute_import(
    file: UploadFile = File(...),
    file_format: str = Form(...),
    site_url: str = Form(...),
    login: str = Form(...),
    password: str = Form(...),
    system_password: str = Form(...),
    selected_format: str = Form(...),
    table_config: str = Form(...),
    reference_lists: str = Form(None)
):
    """
    Execute the import with the uploaded file (Excel only for now)
    Now accepts pre-fetched reference_lists to avoid re-fetching during validation
    """
    try:
        # Parse JSON strings
        selected_format_data = json.loads(selected_format)
        table_config_data = json.loads(table_config)
        reference_lists_data = json.loads(reference_lists) if reference_lists else None
        
        # Save uploaded file temporarily
        upload_dir = Path("/tmp/uploads")
        upload_dir.mkdir(exist_ok=True)
        
        file_path = upload_dir / file.filename
        with open(file_path, "wb") as f:
            content = await file.read()
            f.write(content)
        
        logger.info(f"File uploaded: {file.filename} ({file_format})")
        logger.info(f"Table config: {table_config_data['total_rows']} rows")
        logger.info(f"Selected format: {selected_format_data['name']}")
        
        if file_format != 'excel':
            return {
                "success": False,
                "message": "Seul le format Excel est supporté pour le moment"
            }
        
        # Read Excel file
        logger.info("Lecture du fichier Excel...")
        excel_data = read_excel_file(str(file_path))
        
        if not excel_data['success']:
            return {
                "success": False,
                "message": excel_data['message']
            }
        
        logger.info(f"Excel lu: {len(excel_data['headers'])} colonnes, {len(excel_data['rows'])} lignes")
        
        # Validate key columns
        validation_result = validate_key_columns(excel_data, table_config_data)
        
        if not validation_result['success']:
            return {
                "success": False,
                "message": validation_result['message'],
                "missing_keys": validation_result.get('missing_keys', [])
            }
        
        logger.info("Validation des clés réussie")
        
        # Validate list values
        logger.info("Validation des valeurs de listes...")
        list_validation_result = await validate_list_values(
            excel_data=excel_data,
            table_config=table_config_data,
            site_url=site_url,
            login=login,
            system_password=system_password,
            pre_fetched_lists=reference_lists_data
        )
        
        if not list_validation_result['success']:
            return {
                "success": False,
                "message": list_validation_result['message'],
                "invalid_values": list_validation_result.get('invalid_values', [])
            }
        
        logger.info("Validation des listes réussie")
        
        # Import data to Legisway
        result = await import_to_legisway(
            site_url=site_url,
            login=login,
            password=password,
            selected_format=selected_format_data,
            excel_file_path=str(file_path),
            excel_data=excel_data,
            table_config=table_config_data
        )
        
        return result
        
    except Exception as e:
        logger.error(f"Import error: {str(e)}")
        return {
            "success": False,
            "message": f"Erreur lors de l'import: {str(e)}"
        }

@api_router.get("/import/download-result/{filename}")
async def download_result_file(filename: str):
    """
    Download the result file from Legisway import
    """
    try:
        downloads_dir = Path("/tmp/downloads")
        file_path = downloads_dir / filename
        
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="Fichier non trouvé")
        
        return FileResponse(
            path=str(file_path),
            filename=filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Download error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

def read_excel_file(file_path: str) -> Dict:
    """
    Read Excel file and extract headers and rows
    """
    try:
        workbook = load_workbook(filename=file_path, read_only=True)
        sheet = workbook.active
        
        rows_data = []
        headers = []
        
        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if idx == 0:
                # First row = headers
                headers = [str(cell) if cell is not None else "" for cell in row]
            else:
                # Data rows
                row_data = [str(cell) if cell is not None else "" for cell in row]
                if any(row_data):  # Skip empty rows
                    rows_data.append(row_data)
        
        workbook.close()
        
        return {
            "success": True,
            "headers": headers,
            "rows": rows_data,
            "total_rows": len(rows_data)
        }
        
    except Exception as e:
        logger.error(f"Error reading Excel: {str(e)}")
        return {
            "success": False,
            "message": f"Erreur lecture Excel: {str(e)}",
            "headers": [],
            "rows": [],
            "total_rows": 0
        }

def validate_key_columns(excel_data: Dict, table_config: Dict) -> Dict:
    """
    Validate that key columns (Clé = Oui) are filled in all rows
    """
    try:
        # Extract key fields from table_config
        # table_config has headers and rows
        # Column index 0 = Chemin (path), Column index 1 = Clé (key indicator)
        
        key_fields = []
        
        # Find which fields are marked as keys
        for row in table_config['rows']:
            # row is a dict with 'cells' key
            cells = row.get('cells', []) if isinstance(row, dict) else row.cells
            
            if len(cells) >= 2:
                field_path = cells[0]  # Chemin (path)
                is_key = cells[1]      # Clé (Oui/Non)
                
                if is_key == "Oui":
                    key_fields.append(field_path)
        
        logger.info(f"Champs clés trouvés: {key_fields}")
        
        if not key_fields:
            logger.warning("Aucun champ clé défini dans la configuration")
            return {
                "success": True,
                "message": "Aucun champ clé à valider"
            }
        
        # Map Excel headers to key fields
        excel_headers = excel_data['headers']
        key_column_indices = []
        
        for key_field in key_fields:
            # Try to find matching column in Excel
            found = False
            for idx, header in enumerate(excel_headers):
                if header.strip() == key_field.strip() or key_field.strip() in header.strip():
                    key_column_indices.append((idx, key_field))
                    found = True
                    break
            
            if not found:
                logger.warning(f"Colonne clé '{key_field}' non trouvée dans Excel")
        
        if not key_column_indices:
            return {
                "success": False,
                "message": "Aucune colonne clé trouvée dans le fichier Excel"
            }
        
        logger.info(f"Colonnes clés à valider: {[f[1] for f in key_column_indices]}")
        
        # Validate that key columns are filled in all rows
        missing_keys = []
        
        for row_idx, row in enumerate(excel_data['rows'], start=2):  # Start at 2 (after header)
            for col_idx, key_field in key_column_indices:
                if col_idx < len(row):
                    value = row[col_idx].strip()
                    if not value or value == "":
                        missing_keys.append({
                            "row": row_idx,
                            "column": key_field,
                            "column_index": col_idx + 1
                        })
        
        if missing_keys:
            # Group by column
            missing_by_column = {}
            for missing in missing_keys:
                col = missing['column']
                if col not in missing_by_column:
                    missing_by_column[col] = []
                missing_by_column[col].append(missing['row'])
            
            error_msg = "Colonnes clés manquantes:\n"
            for col, rows in missing_by_column.items():
                rows_str = ", ".join([str(r) for r in rows[:5]])
                if len(rows) > 5:
                    rows_str += f" ... (+{len(rows) - 5} autres)"
                error_msg += f"- '{col}': lignes {rows_str}\n"
            
            return {
                "success": False,
                "message": error_msg,
                "missing_keys": missing_keys
            }
        
        return {
            "success": True,
            "message": f"Toutes les colonnes clés sont remplies ({len(key_column_indices)} colonnes validées)"
        }
        
    except Exception as e:
        logger.error(f"Validation error: {str(e)}")
        return {
            "success": False,
            "message": f"Erreur validation: {str(e)}"
        }

async def validate_list_values(
    excel_data: Dict,
    table_config: Dict,
    site_url: str,
    login: str,
    system_password: str,
    pre_fetched_lists: Optional[Dict] = None
) -> Dict:
    """
    Validate that list values in Excel match allowed values from Legisway
    Now accepts pre_fetched_lists to avoid re-fetching during validation
    """
    try:
        # Extract fields with filters from table_config
        # Column index 0 = Chemin (path), Column index 2 = Filtre (filter)
        
        list_fields = []
        
        for row in table_config['rows']:
            # row is a dict with 'cells' key
            cells = row.get('cells', []) if isinstance(row, dict) else row.cells
            
            if len(cells) >= 3:
                field_path = cells[0]  # Chemin (path)
                filter_value = cells[2]  # Filtre
                
                # Check if filter contains type.name='...'
                if filter_value and "type.name=" in filter_value:
                    # Extract the list type name
                    import re
                    match = re.search(r"type\.name\s*=\s*['\"]([^'\"]+)['\"]", filter_value)
                    if match:
                        list_type = match.group(1)
                        list_fields.append({
                            "field_path": field_path,
                            "list_type": list_type,
                            "filter": filter_value
                        })
        
        logger.info(f"Champs avec listes trouvés: {len(list_fields)}")
        
        if not list_fields:
            return {
                "success": True,
                "message": "Aucune liste à valider"
            }
        
        # Use pre-fetched lists if available, otherwise fetch from Legisway
        if pre_fetched_lists and pre_fetched_lists.get('success') and pre_fetched_lists.get('list_fields'):
            logger.info("Utilisation des listes pré-récupérées")
            # Convert list_fields format to lists format
            list_values_cache = {
                'success': True,
                'lists': {}
            }
            for list_field in pre_fetched_lists['list_fields']:
                list_values_cache['lists'][list_field['list_type']] = list_field['values']
            
            logger.info(f"Listes pré-récupérées: {len(list_values_cache['lists'])}")
            for list_name, list_vals in list_values_cache['lists'].items():
                logger.info(f"  - {list_name}: {len(list_vals)} valeurs")
        else:
            # Fetch list values from Legisway
            logger.info("Récupération des listes depuis Legisway...")
            list_values_cache = await fetch_list_values_from_legisway(
                site_url=site_url,
                login=login,
                system_password=system_password,
                list_types=[field['list_type'] for field in list_fields]
            )
            
            if not list_values_cache['success']:
                return {
                    "success": False,
                    "message": f"Erreur récupération des listes: {list_values_cache['message']}"
                }
            
            logger.info(f"Listes récupérées: {len(list_values_cache['lists'])}")
            for list_name, list_vals in list_values_cache['lists'].items():
                logger.info(f"  - {list_name}: {len(list_vals)} valeurs")
        
        # Map Excel headers to list fields
        excel_headers = excel_data['headers']
        logger.info(f"En-têtes Excel: {excel_headers}")
        list_column_indices = []
        
        for list_field in list_fields:
            field_path = list_field['field_path']
            list_type = list_field['list_type']
            logger.info(f"Recherche colonne pour: {field_path} (type: {list_type})")
            
            # Find matching column in Excel
            # Try multiple matching strategies
            for idx, header in enumerate(excel_headers):
                header_clean = header.strip().lower()
                field_clean = field_path.strip().lower()
                
                # Strategy 1: Exact match
                if header_clean == field_clean:
                    allowed_vals = list_values_cache['lists'].get(list_type, [])
                    logger.info(f"  -> Trouvé (exact) à l'index {idx}, {len(allowed_vals)} valeurs autorisées")
                    list_column_indices.append({
                        "col_idx": idx,
                        "field_path": field_path,
                        "list_type": list_type,
                        "allowed_values": allowed_vals
                    })
                    break
                
                # Strategy 2: Field path in header
                if field_clean in header_clean:
                    allowed_vals = list_values_cache['lists'].get(list_type, [])
                    logger.info(f"  -> Trouvé (contenu) à l'index {idx}, {len(allowed_vals)} valeurs autorisées")
                    list_column_indices.append({
                        "col_idx": idx,
                        "field_path": field_path,
                        "list_type": list_type,
                        "allowed_values": allowed_vals
                    })
                    break
                
                # Strategy 3: Extract last part after dot (e.g., "title.fr" from "internalExternal.title.fr")
                # and check if it's in the header
                if '.' in field_clean:
                    field_suffix = field_clean.split('.')[-1]  # Get "fr"
                    field_last_two = '.'.join(field_clean.split('.')[-2:])  # Get "title.fr"
                    
                    if field_last_two in header_clean or field_suffix in header_clean:
                        # Additional check: verify the header is related to the list type
                        # by checking if list type name parts are in header
                        list_type_clean = list_type.lower()
                        # Extract base name: "internalExternalList" -> "internal", "external"
                        if 'internal' in field_clean and 'internal' in header_clean:
                            allowed_vals = list_values_cache['lists'].get(list_type, [])
                            logger.info(f"  -> Trouvé (suffix+context) à l'index {idx}, {len(allowed_vals)} valeurs autorisées")
                            list_column_indices.append({
                                "col_idx": idx,
                                "field_path": field_path,
                                "list_type": list_type,
                                "allowed_values": allowed_vals
                            })
                            break
                        elif 'civility' in field_clean and 'civilité' in header_clean:
                            allowed_vals = list_values_cache['lists'].get(list_type, [])
                            logger.info(f"  -> Trouvé (civility) à l'index {idx}, {len(allowed_vals)} valeurs autorisées")
                            list_column_indices.append({
                                "col_idx": idx,
                                "field_path": field_path,
                                "list_type": list_type,
                                "allowed_values": allowed_vals
                            })
                            break
                        elif 'function' in field_clean and 'fonction' in header_clean:
                            allowed_vals = list_values_cache['lists'].get(list_type, [])
                            logger.info(f"  -> Trouvé (function) à l'index {idx}, {len(allowed_vals)} valeurs autorisées")
                            list_column_indices.append({
                                "col_idx": idx,
                                "field_path": field_path,
                                "list_type": list_type,
                                "allowed_values": allowed_vals
                            })
                            break
                        elif 'department' in field_clean and 'direction' in header_clean:
                            allowed_vals = list_values_cache['lists'].get(list_type, [])
                            logger.info(f"  -> Trouvé (department) à l'index {idx}, {len(allowed_vals)} valeurs autorisées")
                            list_column_indices.append({
                                "col_idx": idx,
                                "field_path": field_path,
                                "list_type": list_type,
                                "allowed_values": allowed_vals
                            })
                            break
                        elif 'company' in field_clean and 'société' in header_clean:
                            allowed_vals = list_values_cache['lists'].get(list_type, [])
                            logger.info(f"  -> Trouvé (company) à l'index {idx}, {len(allowed_vals)} valeurs autorisées")
                            list_column_indices.append({
                                "col_idx": idx,
                                "field_path": field_path,
                                "list_type": list_type,
                                "allowed_values": allowed_vals
                            })
                            break
        
        if not list_column_indices:
            logger.warning("Aucune colonne de liste trouvée dans Excel!")
            return {
                "success": True,
                "message": "Aucune colonne de liste trouvée dans Excel"
            }
        
        logger.info(f"Colonnes de listes à valider: {len(list_column_indices)}")
        
        # Validate values
        invalid_values = []
        values_checked = 0
        
        for row_idx, row in enumerate(excel_data['rows'], start=2):  # Start at 2 (after header)
            for list_col in list_column_indices:
                col_idx = list_col['col_idx']
                
                if col_idx < len(row):
                    value = row[col_idx].strip()
                    
                    # Empty values are allowed
                    if value and value != "":
                        values_checked += 1
                        allowed_values = list_col['allowed_values']
                        
                        logger.info(f"Validation ligne {row_idx}, colonne {list_col['field_path']}: valeur='{value}', nb_autorisées={len(allowed_values)}")
                        
                        # Check if value is in allowed list
                        if value not in allowed_values:
                            logger.warning(f"  -> INVALIDE: '{value}' non trouvée dans {len(allowed_values)} valeurs autorisées")
                            invalid_values.append({
                                "row": row_idx,
                                "column": list_col['field_path'],
                                "value": value,
                                "list_type": list_col['list_type'],
                                "allowed_values": allowed_values[:10]  # Show first 10 for error message
                            })
                        else:
                            logger.info(f"  -> VALIDE")
        
        logger.info(f"Validation terminée: {values_checked} valeurs vérifiées, {len(invalid_values)} invalides")
        
        if invalid_values:
            # Group by column
            invalid_by_column = {}
            for invalid in invalid_values:
                col = invalid['column']
                if col not in invalid_by_column:
                    invalid_by_column[col] = []
                invalid_by_column[col].append({
                    "row": invalid['row'],
                    "value": invalid['value']
                })
            
            error_msg = "Valeurs invalides dans les listes:\n"
            for col, errors in invalid_by_column.items():
                error_msg += f"\n- Colonne '{col}':\n"
                for err in errors[:5]:  # Show first 5
                    error_msg += f"  Ligne {err['row']}: '{err['value']}' (invalide)\n"
                if len(errors) > 5:
                    error_msg += f"  ... (+{len(errors) - 5} autres valeurs invalides)\n"
            
            # Show allowed values for first error
            if invalid_values:
                first_invalid = invalid_values[0]
                error_msg += f"\nValeurs autorisées pour '{first_invalid['list_type']}': "
                error_msg += ", ".join([f"'{v}'" for v in first_invalid['allowed_values'][:10]])
                if len(first_invalid['allowed_values']) > 10:
                    error_msg += f" ... (+{len(first_invalid['allowed_values']) - 10} autres)"
            
            return {
                "success": False,
                "message": error_msg,
                "invalid_values": invalid_values
            }
        
        return {
            "success": True,
            "message": f"Toutes les valeurs de listes sont valides ({len(list_column_indices)} colonnes validées)"
        }
        
    except Exception as e:
        logger.error(f"List validation error: {str(e)}")
        return {
            "success": False,
            "message": f"Erreur validation listes: {str(e)}"
        }

async def fetch_list_values_from_legisway(
    site_url: str,
    login: str,
    system_password: str,
    list_types: List[str]
) -> Dict:
    """
    Fetch allowed values for reference lists from Legisway using REST API
    """
    try:
        # Get base URL from site_url
        from urllib.parse import urlparse
        parsed = urlparse(site_url)
        base_url = f"{parsed.scheme}://{parsed.netloc}"
        
        async with httpx.AsyncClient(verify=False, timeout=60.0) as client:
            # Step 1: Authenticate and get JWT token using system password
            logger.info(f"Authentication système à l'API Legisway...")
            auth_url = f"{base_url}/resource/api/v1/auth/system"
            
            # System authentication uses only system password (not user password)
            auth_payload = {
                "password": system_password,
                "languageCode": "fr"
            }
            
            logger.info(f"Auth URL: {auth_url}")
            
            try:
                # Send as form-data (application/x-www-form-urlencoded)
                auth_response = await client.post(
                    auth_url,
                    data=auth_payload,
                    headers={
                        "Accept": "application/json"
                    }
                )
                
                logger.info(f"Auth response status: {auth_response.status_code}")
                
                if auth_response.status_code != 200:
                    logger.error(f"Authentication failed: {auth_response.status_code}, body: {auth_response.text[:200]}")
                    return {
                        "success": False,
                        "message": f"Échec authentification API: {auth_response.status_code} - {auth_response.text[:100]}",
                        "lists": {}
                    }
                
                # Remove quotes from JWT token
                jwt_token = auth_response.text.strip('"')
                logger.info(f"Token JWT obtenu (length: {len(jwt_token)})")
                
            except Exception as e:
                logger.error(f"Auth request error: {str(e)}")
                return {
                    "success": False,
                    "message": f"Erreur requête auth: {str(e)}",
                    "lists": {}
                }
            
            # Step 2: For each list type, fetch field definitions
            lists = {}
            
            for list_type in set(list_types):  # Use set to avoid duplicates
                try:
                    logger.info(f"Récupération de la liste: {list_type}")
                    
                    # Try multiple approaches to get list values
                    values = []
                    search_response = None
                    
                    # Approach 1: Try GET /search/{list_type}
                    try:
                        search_url = f"{base_url}/resource/api/v1/search/{list_type}?offset=0&limit=1000"
                        logger.info(f"Tentative GET: {search_url}")
                        search_response = await client.get(
                            search_url,
                            headers={
                                "Accept": "application/json",
                                "Authorization": f"Bearer {jwt_token}"
                            }
                        )
                        logger.info(f"GET response: {search_response.status_code}")
                    except Exception as e:
                        logger.warning(f"GET /search failed: {str(e)}")
                    
                    # Approach 2: Try GET /{list_type}
                    if not search_response or search_response.status_code != 200:
                        try:
                            direct_url = f"{base_url}/resource/api/v1/{list_type}?offset=0&limit=1000"
                            logger.info(f"Tentative GET direct: {direct_url}")
                            search_response = await client.get(
                                direct_url,
                                headers={
                                    "Accept": "application/json",
                                    "Authorization": f"Bearer {jwt_token}"
                                }
                            )
                            logger.info(f"GET direct response: {search_response.status_code}")
                        except Exception as e:
                            logger.warning(f"GET direct failed: {str(e)}")
                    
                    if search_response and search_response.status_code == 200:
                        search_data = search_response.json()
                        
                        # Check if response is a list (Legisway format) or dict with 'data' key
                        if isinstance(search_data, list):
                            logger.info(f"API retourné une liste directe avec {len(search_data)} items")
                            # Process list directly - Legisway returns list with _title field
                            for item in search_data:
                                if isinstance(item, dict):
                                    # Try different field names
                                    if '_title' in item:
                                        values.append(item['_title'])
                                    elif 'title' in item:
                                        if isinstance(item['title'], dict) and 'fr' in item['title']:
                                            values.append(item['title']['fr'])
                                        else:
                                            values.append(str(item['title']))
                                    elif 'name' in item:
                                        values.append(item['name'])
                            
                            logger.info(f"Liste {list_type}: {len(values)} valeurs extraites")
                            if len(values) > 0:
                                logger.info(f"Exemples: {values[:5]}")
                        
                        elif isinstance(search_data, dict) and 'data' in search_data:
                            logger.info(f"API retourné un dict avec clé 'data' ({len(search_data['data'])} items)")
                            for item in search_data['data']:
                                # Try to get the title or name
                                if '_title' in item:
                                    values.append(item['_title'])
                                elif 'title' in item and isinstance(item['title'], dict) and 'fr' in item['title']:
                                    values.append(item['title']['fr'])
                                elif 'name' in item:
                                    values.append(item['name'])
                            
                            logger.info(f"Liste {list_type}: {len(values)} valeurs récupérées")
                            if len(values) > 0:
                                logger.info(f"Exemples: {values[:5]}")
                        
                        else:
                            logger.warning(f"Format de réponse inattendu pour {list_type}")
                            logger.info(f"Type: {type(search_data)}, Contenu: {str(search_data)[:500]}")
                    else:
                        logger.warning(f"Erreur récupération {list_type}: {search_response.status_code}")
                        logger.info(f"Réponse erreur: {search_response.text[:500]}")
                    
                    lists[list_type] = list(set(values))  # Remove duplicates
                    logger.info(f"Liste {list_type}: {len(lists[list_type])} valeurs uniques")
                        
                except Exception as e:
                    logger.error(f"Erreur pour {list_type}: {str(e)}")
                    lists[list_type] = []
            
            return {
                "success": True,
                "lists": lists
            }
                
    except Exception as e:
        logger.error(f"Fetch list values error: {str(e)}")
        return {
            "success": False,
            "message": f"Erreur récupération listes: {str(e)}",
            "lists": {}
        }

async def click_login_button(page, browser=None):
    """
    Helper function to click the login button with multiple selector attempts
    Returns True if successful, False otherwise
    """
    selectors_to_try = [
        'button:has-text("Connexion")',
        'button.button-large:has-text("Connexion")',
        'button[data-kind="button"]:has-text("Connexion")',
        'button.button-large-with-label',
        'button[type="button"]:has-text("Connexion")'
    ]
    
    for selector in selectors_to_try:
        try:
            await page.wait_for_selector(selector, timeout=3000)
            logger.info(f"Bouton connexion trouvé avec: {selector}")
            await asyncio.sleep(0.5)
            await page.click(selector, timeout=5000)
            return True
        except Exception as e:
            continue
    
    logger.error("Aucun bouton de connexion trouvé")
    return False

async def import_to_legisway(
    site_url: str,
    login: str,
    password: str,
    selected_format: Dict,
    excel_file_path: str,
    excel_data: Dict,
    table_config: Dict
) -> Dict:
    """
    Import Excel data to Legisway using Playwright automation with rollback test option
    Returns the result file for user download
    """
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True, args=['--no-sandbox', '--disable-setuid-sandbox'])
            context = await browser.new_context(
                ignore_https_errors=True,
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                accept_downloads=True
            )
            page = await context.new_page()
            
            try:
                # Step 1: Navigate and login
                logger.info("Connexion à Legisway pour import...")
                await page.goto(site_url, timeout=30000, wait_until="load")
                await asyncio.sleep(2)
                
                # Check if already logged in
                already_logged_in = await page.query_selector('.icon-user')
                
                if not already_logged_in:
                    logger.info("Connexion requise...")
                    await page.fill('input[name="j_username"]', login, timeout=5000)
                    await asyncio.sleep(0.5)
                    await page.fill('input[name="j_password"]', password, timeout=5000)
                    await asyncio.sleep(1)
                    
                    await page.wait_for_selector('button:has-text("Connexion")', timeout=10000)
                    await asyncio.sleep(0.5)
                    await page.click('button:has-text("Connexion")', timeout=5000)
                    
                    await page.wait_for_load_state("load", timeout=30000)
                    await asyncio.sleep(3)
                else:
                    logger.info("Déjà connecté")
                
                # Step 2: Navigate to Import section
                logger.info("Navigation vers Import de données...")
                await page.wait_for_selector('.icon-user', timeout=10000)
                await page.click('.icon-user', timeout=5000)
                await asyncio.sleep(1)
                
                await page.wait_for_selector('button[mat-menu-item]', timeout=10000)
                await page.click('button.user-menu-item:has-text("Administration")', timeout=5000)
                await page.wait_for_load_state("load", timeout=30000)
                await asyncio.sleep(3)
                
                await page.wait_for_selector('a:has-text("Import de données")', timeout=10000)
                await page.click('a:has-text("Import de données")', timeout=10000)
                await page.wait_for_load_state("load", timeout=30000)
                await asyncio.sleep(3)
                
                # Step 3: Select format in combobox
                logger.info(f"Sélection du format: {selected_format['name']}")
                await page.wait_for_selector('kendo-combobox input.k-input-inner', timeout=10000)
                await page.click('kendo-combobox input.k-input-inner')
                await asyncio.sleep(1)
                await page.fill('kendo-combobox input.k-input-inner', selected_format['name'])
                await asyncio.sleep(1)
                await page.keyboard.press('Enter')
                await asyncio.sleep(2)
                
                # Step 4: Upload Excel file
                logger.info("Upload du fichier Excel...")
                # Find the file input (usually hidden in dropzone)
                file_input = await page.query_selector('input[type="file"]')
                if file_input:
                    await file_input.set_input_files(excel_file_path)
                    logger.info("Fichier uploadé via input")
                else:
                    logger.warning("Input file non trouvé, essai avec dropzone")
                    # Alternative: try to trigger file input via dropzone click
                    await page.click('.dropzone')
                    await asyncio.sleep(1)
                    file_input = await page.query_selector('input[type="file"]')
                    if file_input:
                        await file_input.set_input_files(excel_file_path)
                
                await asyncio.sleep(2)
                
                # Step 5: Open Advanced Options and select rollback
                logger.info("Configuration du mode rollback...")
                await page.wait_for_selector('button[aria-label*="Options Avancées"]', timeout=10000)
                await page.click('button[aria-label*="Options Avancées"]')
                await asyncio.sleep(1)
                
                # Select rollback radio button
                await page.wait_for_selector('mat-radio-button[value="with.rollback"] input', timeout=10000)
                await page.click('mat-radio-button[value="with.rollback"] input')
                await asyncio.sleep(1)
                logger.info("Mode rollback activé")
                
                # Step 6: Click Import button
                logger.info("Lancement de l'import...")
                await page.wait_for_selector('button:has-text("Importer"):not([disabled])', timeout=10000)
                await page.click('button:has-text("Importer")')
                await asyncio.sleep(2)
                
                # Step 7: Wait for progress bar to complete
                logger.info("Attente de la fin de l'import...")
                # Wait for progress bar to appear
                await page.wait_for_selector('mat-progress-bar', timeout=10000)
                await asyncio.sleep(2)
                
                # Wait for progress to reach 100% or import to finish
                # Check for completion message (success or failure)
                max_wait_time = 3600  # 1 hour max
                check_interval = 5
                elapsed = 0
                
                while elapsed < max_wait_time:
                    # Check if import finished (look for result file link or completion message)
                    result_file = await page.query_selector('a[href*="result_file"]')
                    completion_message = await page.query_selector('div:has-text("Import terminé"), div:has-text("Échec de l\'import")')
                    
                    if result_file or completion_message:
                        logger.info("Import terminé!")
                        break
                    
                    # Check progress percentage
                    progress_label = await page.query_selector('mat-progress-bar + label')
                    if progress_label:
                        progress_text = await progress_label.text_content()
                        logger.info(f"Progression: {progress_text}")
                    
                    await asyncio.sleep(check_interval)
                    elapsed += check_interval
                
                if elapsed >= max_wait_time:
                    await browser.close()
                    return {
                        "success": False,
                        "message": "Timeout: l'import a pris plus d'1 heure"
                    }
                
                await asyncio.sleep(2)
                
                # Step 8: Get result file
                logger.info("Récupération du fichier de résultat...")
                result_file_link = await page.query_selector('a[href*="result_file"]')
                
                if result_file_link:
                    result_file_name = await result_file_link.text_content()
                    result_file_href = await result_file_link.get_attribute('href')
                    
                    logger.info(f"Fichier de résultat: {result_file_name}")
                    
                    # Download the result file
                    async with page.expect_download() as download_info:
                        await result_file_link.click()
                    download = await download_info.value
                    
                    # Save to downloads directory
                    downloads_dir = Path("/tmp/downloads")
                    downloads_dir.mkdir(exist_ok=True)
                    result_file_path = downloads_dir / result_file_name.strip()
                    
                    await download.save_as(str(result_file_path))
                    logger.info(f"Fichier de résultat sauvegardé: {result_file_path}")
                    
                    await browser.close()
                    
                    return {
                        "success": True,
                        "message": f"Import terminé avec rollback: {excel_data['total_rows']} lignes traitées",
                        "rows_imported": excel_data['total_rows'],
                        "result_file_path": str(result_file_path),
                        "result_file_name": result_file_name.strip()
                    }
                else:
                    # No result file found - check for error message
                    error_message = await page.query_selector('div:has-text("Échec")')
                    if error_message:
                        error_text = await error_message.text_content()
                        await browser.close()
                        return {
                            "success": False,
                            "message": f"Import échoué: {error_text}"
                        }
                    else:
                        await browser.close()
                        return {
                            "success": True,
                            "message": f"Import terminé (pas de fichier de résultat disponible)"
                        }
                
            except Exception as e:
                logger.error(f"Erreur pendant l'import: {str(e)}")
                await browser.close()
                raise e
                
    except Exception as e:
        logger.error(f"Legisway import error: {str(e)}")
        return {
            "success": False,
            "message": f"Erreur import Legisway: {str(e)}"
        }

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()